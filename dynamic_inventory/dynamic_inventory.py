#!/usr/bin/env python3

import sys
import json
import os
from openpyxl import load_workbook

# Ruta dinámica del archivo Excel basada en la ubicación del script
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_file = os.path.join(script_dir, 'inventory_data.xlsx')

def parse_excel(file):
    try:
        # Cargar el archivo Excel
        wb = load_workbook(file)
        sheet = wb.active  # Seleccionar la hoja activa
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo '{file}'. Verifica que esté en la ruta correcta.")
        sys.exit(1)
    except Exception as e:
        print(f"Error al leer el archivo Excel: {e}")
        sys.exit(1)

    # Estructura del inventario
    inventory = {"_meta": {"hostvars": {}}}

    try:
        # Iterar por las filas del archivo Excel (omitiendo la cabecera)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Extraer columnas según el orden: host, grupo, IP, usuario, descripción, intérprete Python
            host, group, ip, user, description, python_interpreter = row

            # Validar que las columnas obligatorias tengan datos
            if not host or not group or not ip or not user:
                print(f"Error: La fila con valores {row} tiene datos faltantes. Verifica el archivo Excel.")
                sys.exit(1)

            # Agregar host al grupo correspondiente
            if group not in inventory:
                inventory[group] = {"hosts": [], "vars": {}}
            inventory[group]["hosts"].append(host)

            # Agregar variables del host
            inventory["_meta"]["hostvars"][host] = {
                "ansible_host": ip,
                "ansible_user": user,
                "description": description or "No description provided"
            }

            # Agregar el campo ansible_python_interpreter si está definido en el Excel
            if python_interpreter:
                inventory["_meta"]["hostvars"][host]["ansible_python_interpreter"] = python_interpreter

    except Exception as e:
        print(f"Error al procesar el archivo Excel: {e}")
        sys.exit(1)

    return inventory

def main():
    # Revisar argumentos pasados al script
    if len(sys.argv) == 2 and sys.argv[1] == '--list':
        try:
            # Generar inventario dinámico
            inventory = parse_excel(excel_file)
            print(json.dumps(inventory, indent=2))
        except Exception as e:
            print(f"Error al generar el inventario dinámico: {e}")
            sys.exit(1)
    elif len(sys.argv) == 3 and sys.argv[1] == '--host':
        # Devuelve las variables específicas de un host (opcional)
        host = sys.argv[2]
        try:
            inventory = parse_excel(excel_file)
            host_vars = inventory["_meta"]["hostvars"].get(host, {})
            print(json.dumps(host_vars, indent=2))
        except Exception as e:
            print(f"Error al obtener variables del host: {e}")
            sys.exit(1)
    else:
        print("Uso: --list | --host <nombre_host>")
        sys.exit(1)

if __name__ == '__main__':
    main()

